import pandas as pd
from openpyxl import load_workbook
from ai_processor import HeaderNavItem, FooterNavItem, WebsiteAsset, ContentFramework
import os

def parse_example_xlsx(path: str) -> ContentFramework:
    if not os.path.exists(path):
        return ContentFramework(header_nav=[], footer_nav=[], website_assets=[], cta_strategy="")
    wb = load_workbook(path)
    header_nav = []
    footer_nav = []
    website_assets = []
    cta_strategy = ""

    # Read sheets by expected names
    for sheet_name in ["header navigation content", "footer navigation content", "website assets"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        data = []
        # Iterate rows, assume headers are in row 2 (based on styling startrow=1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
        df = pd.DataFrame(data, columns=[cell.value for cell in ws[1] if cell.value is not None])
        # Map columns to model fields
        for _, row in df.iterrows():
            if sheet_name == "header navigation content":
                header_nav.append(HeaderNavItem(
                    main_nav=row.get("MAIN NAV. ITEM/LAUNCH POINT", ""),
                    dropdown=row.get("DROPDOWN/NEXT STOP", ""),
                    final_destination=row.get("FINAL DESTINATION", ""),
                    page_type=row.get("PAGE TYPE", ""),
                    page_description=row.get("PAGE DESCRIPTION", ""),
                    key_sections=row.get("KEY SECTIONS/FEATURES", ""),
                    content_type=row.get("CONTENT TYPE", ""),
                    content_link=row.get("🔗 CONTENT LINK", ""),
                    status=row.get("📝 STATUS", ""),
                    client_notes=row.get("💬 CLIENT NOTES", "")
                ))
            elif sheet_name == "footer navigation content":
                footer_nav.append(FooterNavItem(
                    menu_title=row.get("FOOTER MENU TITLE", ""),
                    nested_items=row.get("NESTED MENU ITEMS", ""),
                    page_type=row.get("PAGE TYPE", ""),
                    page_description=row.get("PAGE DESCRIPTION", ""),
                    key_sections=row.get("KEY SECTIONS/FEATURES", ""),
                    content_type=row.get("CONTENT TYPE", ""),
                    content_link=row.get("🔗 CONTENT LINK", ""),
                    status=row.get("📝 STATUS", ""),
                    client_notes=row.get("💬 CLIENT NOTES", "")
                ))
            elif sheet_name == "website assets":
                website_assets.append(WebsiteAsset(
                    asset_required=row.get("ASSETS REQUIRED", ""),
                    description=row.get("DESCRIPTION", ""),
                    content_type=row.get("CONTENT TYPE", ""),
                    content_link=row.get("🔗 CONTENT LINK", ""),
                    status=row.get("📝 STATUS", ""),
                    client_notes=row.get("💬 CLIENT NOTES", "")
                ))
            # Look for CTA strategy in a separate sheet or cell
            # Fallback: assume it's in a sheet named differently or in a specific cell
            for sheet in wb.sheetnames:
                if "cta" in sheet.lower() or "strategy" in sheet.lower():
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        if isinstance(row, str) and "call to action" in row.lower():
                            cta_strategy = row
                            break

    return ContentFramework(
        header_nav=header_nav,
        footer_nav=footer_nav,
        website_assets=website_assets,
        cta_strategy=cta_strategy
    )

def match_example_by_input(input_text: str, engine) -> ContentFramework | None:
    """
    Find best-matching example by input similarity using the validation engine embeddings.
    Returns the parsed ContentFramework if a high-confidence match is found.
    """
    examples_dir = "sample_outputs"
    if not os.path.exists(examples_dir):
        return None
    example_files = [f for f in os.listdir(examples_dir) if f.endswith('.xlsx')]
    best_match = None
    best_score = -1
    for example_file in example_files:
        example_path = os.path.join(examples_dir, example_file)
        try:
            example_framework = parse_example_xlsx(example_path)
            # Combine all text from the example for comparison
            example_text = "\n".join([
                example_framework.cta_strategy or "",
                " ".join([item.main_nav or "" for item in example_framework.header_nav]),
                " ".join([item.menu_title or "" for item in example_framework.footer_nav]),
                " ".join([item.asset_required or "" for item in example_framework.website_assets])
            ])
            emb = engine.get_embedding(example_text)
            input_emb = engine.get_embedding(input_text)
            # Simple cosine similarity
            import numpy as np
            score = float(np.dot(input_emb.flatten(), emb.flatten()) / (
                np.linalg.norm(input_emb) * np.linalg.norm(emb) + 1e-8))
            if score > best_score:
                best_score = score
                best_match = example_framework
        except Exception:
            continue
    # Threshold: only return match if similarity is reasonably high
    if best_score > 0.85:
        return best_match
    return None
