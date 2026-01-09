import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def _clean_val(val):
    if isinstance(val, list):
        return "\n".join([f"• {str(i)}" for i in val])
    return str(val) if val is not None else ""

def _get(item, key, default=""):
    if isinstance(item, dict):
        return item.get(key, default)
    return getattr(item, key, default)

def _apply_standard_styling(sheet, df, title_text):
    # Colors
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    title_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, name="Arial", size=12)
    
    cyan_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
    purple_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
    
    border = Border(
        left=Side(style='thin', color="DDDDDD"),
        right=Side(style='thin', color="DDDDDD"),
        top=Side(style='thin', color="DDDDDD"),
        bottom=Side(style='thin', color="DDDDDD")
    )

    # Row 1: Title
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = title_text
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 25

    # Row 2: Headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = column_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[2].height = 40

    # Data Styling
    for row_num in range(3, len(df) + 3):
        sheet.row_dimensions[row_num].height = 60 # Generous height like in image
        for col_num in range(1, len(df.columns) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            cell.border = border
            
            # Apply specific column colors for Header/Footer structure
            if "NAV" in title_text or "FOOTER" in title_text:
                if col_num == 1 and cell.value:
                    cell.fill = cyan_fill
                elif col_num == 2 and cell.value:
                    cell.fill = purple_fill
                elif col_num == 3 and cell.value:
                    cell.fill = yellow_fill

    # Column Widths
    for i, col in enumerate(df.columns):
        sheet.column_dimensions[get_column_letter(i+1)].width = 25

def get_header_nav_excel(framework):
    output = BytesIO()
    data = []
    for item in framework.header_nav:
        data.append({
            'MAIN NAV. ITEM/LAUNCH POINT': _clean_val(_get(item, 'main_nav', '')),
            'DROPDOWN/NEXT STOP': _clean_val(_get(item, 'dropdown', '')),
            'FINAL DESTINATION': _clean_val(_get(item, 'final_destination', '')),
            'PAGE TYPE': _clean_val(_get(item, 'page_type', '')),
            'PAGE DESCRIPTION': _clean_val(_get(item, 'page_description', '')),
            'KEY SECTIONS/FEATURES': _clean_val(_get(item, 'key_sections', '')),
            'CONTENT TYPE': _clean_val(_get(item, 'content_type', '')),
            '🔗 CONTENT LINK': _clean_val(_get(item, 'content_link', '')),
            '📝 STATUS': _clean_val(_get(item, 'status', '⏳ Not Started')),
            '💬 CLIENT NOTES': _clean_val(_get(item, 'client_notes', ''))
        })
    df = pd.DataFrame(data)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1)
        _apply_standard_styling(writer.book.active, df, "1️⃣ HEADER NAVIGATION ITEMS")
    return output.getvalue()

def get_footer_nav_excel(framework):
    output = BytesIO()
    data = []
    for item in framework.footer_nav:
        data.append({
            'FOOTER MENU TITLE': _clean_val(_get(item, 'menu_title', '')),
            'NESTED MENU ITEMS': _clean_val(_get(item, 'nested_items', '')),
            'PAGE TYPE': _clean_val(_get(item, 'page_type', '')),
            'PAGE DESCRIPTION': _clean_val(_get(item, 'page_description', '')),
            'KEY SECTIONS/FEATURES': _clean_val(_get(item, 'key_sections', '')),
            'CONTENT TYPE': _clean_val(_get(item, 'content_type', '')),
            '🔗 CONTENT LINK': _clean_val(_get(item, 'content_link', '')),
            '📝 STATUS': _clean_val(_get(item, 'status', '⏳ Not Started')),
            '💬 CLIENT NOTES': _clean_val(_get(item, 'client_notes', ''))
        })
    df = pd.DataFrame(data)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1)
        _apply_standard_styling(writer.book.active, df, "2️⃣ FOOTER NAVIGATION ITEMS")
    return output.getvalue()

def get_website_assets_excel(framework):
    output = BytesIO()
    data = []
    for item in framework.website_assets:
        data.append({
            'ASSETS REQUIRED': _clean_val(_get(item, 'asset_required', '')),
            'DESCRIPTION': _clean_val(_get(item, 'description', '')),
            'CONTENT TYPE': _clean_val(_get(item, 'content_type', '')),
            '🔗 CONTENT LINK': _clean_val(_get(item, 'content_link', '')),
            '📝 STATUS': _clean_val(_get(item, 'status', '⏳ Not Started')),
            '💬 CLIENT NOTES': _clean_val(_get(item, 'client_notes', ''))
        })
    df = pd.DataFrame(data)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1)
        _apply_standard_styling(writer.book.active, df, "3️⃣ WEBSITE ASSETS")
    return output.getvalue()

def framework_to_excel(framework):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Header
        h_data = []
        for item in framework.header_nav:
            h_data.append({
                'MAIN NAV. ITEM/LAUNCH POINT': _clean_val(_get(item, 'main_nav', '')),
                'DROPDOWN/NEXT STOP': _clean_val(_get(item, 'dropdown', '')),
                'FINAL DESTINATION': _clean_val(_get(item, 'final_destination', '')),
                'PAGE TYPE': _clean_val(_get(item, 'page_type', '')),
                'PAGE DESCRIPTION': _clean_val(_get(item, 'page_description', '')),
                'KEY SECTIONS/FEATURES': _clean_val(_get(item, 'key_sections', '')),
                'CONTENT TYPE': _clean_val(_get(item, 'content_type', '')),
                '🔗 CONTENT LINK': _clean_val(_get(item, 'content_link', '')),
                '📝 STATUS': _clean_val(_get(item, 'status', '⏳ Not Started')),
                '💬 CLIENT NOTES': _clean_val(_get(item, 'client_notes', ''))
            })
        df_h = pd.DataFrame(h_data)
        df_h.to_excel(writer, sheet_name='header navigation content', index=False, startrow=1)
        _apply_standard_styling(writer.book['header navigation content'], df_h, "1️⃣ HEADER NAVIGATION ITEMS")

        # Footer
        f_data = []
        for item in framework.footer_nav:
            f_data.append({
                'FOOTER MENU TITLE': _clean_val(_get(item, 'menu_title', '')),
                'NESTED MENU ITEMS': _clean_val(_get(item, 'nested_items', '')),
                'PAGE TYPE': _clean_val(_get(item, 'page_type', '')),
                'PAGE DESCRIPTION': _clean_val(_get(item, 'page_description', '')),
                'KEY SECTIONS/FEATURES': _clean_val(_get(item, 'key_sections', '')),
                'CONTENT TYPE': _clean_val(_get(item, 'content_type', '')),
                '🔗 CONTENT LINK': _clean_val(_get(item, 'content_link', '')),
                '📝 STATUS': _clean_val(_get(item, 'status', '⏳ Not Started')),
                '💬 CLIENT NOTES': _clean_val(_get(item, 'client_notes', ''))
            })
        df_f = pd.DataFrame(f_data)
        df_f.to_excel(writer, sheet_name='footer navigation content', index=False, startrow=1)
        _apply_standard_styling(writer.book['footer navigation content'], df_f, "2️⃣ FOOTER NAVIGATION ITEMS")

        # Assets
        a_data = []
        for item in framework.website_assets:
            a_data.append({
                'ASSETS REQUIRED': _clean_val(_get(item, 'asset_required', '')),
                'DESCRIPTION': _clean_val(_get(item, 'description', '')),
                'CONTENT TYPE': _clean_val(_get(item, 'content_type', '')),
                '🔗 CONTENT LINK': _clean_val(_get(item, 'content_link', '')),
                '📝 STATUS': _clean_val(_get(item, 'status', '⏳ Not Started')),
                '💬 CLIENT NOTES': _clean_val(_get(item, 'client_notes', ''))
            })
        df_a = pd.DataFrame(a_data)
        df_a.to_excel(writer, sheet_name='website assets', index=False, startrow=1)
        _apply_standard_styling(writer.book['website assets'], df_a, "3️⃣ WEBSITE ASSETS")

    output.seek(0)
    return output.getvalue()

def get_blank_framework_excel():
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Define columns
        h_cols = ['MAIN NAV. ITEM/LAUNCH POINT', 'DROPDOWN/NEXT STOP', 'FINAL DESTINATION', 'PAGE TYPE', 'PAGE DESCRIPTION', 'KEY SECTIONS/FEATURES', 'CONTENT TYPE', '🔗 CONTENT LINK', '📝 STATUS', '💬 CLIENT NOTES']
        f_cols = ['FOOTER MENU TITLE', 'NESTED MENU ITEMS', 'PAGE TYPE', 'PAGE DESCRIPTION', 'KEY SECTIONS/FEATURES', 'CONTENT TYPE', '🔗 CONTENT LINK', '📝 STATUS', '💬 CLIENT NOTES']
        a_cols = ['ASSETS REQUIRED', 'DESCRIPTION', 'CONTENT TYPE', '🔗 CONTENT LINK', '📝 STATUS', '💬 CLIENT NOTES']

        # Create empty DataFrames
        df_h = pd.DataFrame(columns=h_cols)
        df_f = pd.DataFrame(columns=f_cols)
        df_a = pd.DataFrame(columns=a_cols)

        # Write to sheets
        df_h.to_excel(writer, sheet_name='header navigation content', index=False, startrow=1)
        _apply_standard_styling(writer.book['header navigation content'], df_h, "1️⃣ HEADER NAVIGATION ITEMS")

        df_f.to_excel(writer, sheet_name='footer navigation content', index=False, startrow=1)
        _apply_standard_styling(writer.book['footer navigation content'], df_f, "2️⃣ FOOTER NAVIGATION ITEMS")

        df_a.to_excel(writer, sheet_name='website assets', index=False, startrow=1)
        _apply_standard_styling(writer.book['website assets'], df_a, "3️⃣ WEBSITE ASSETS")

    output.seek(0)
    return output.getvalue()

def scope_to_excel(scope):
    output = BytesIO()
    overview_data = [
        ['Field', 'Description'],
        ['Project Title', scope.project_title],
        ['Objectives', "\n".join(scope.objectives)],
        ['In-Scope', "\n".join(scope.scope_in)],
        ['Out-of-Scope', "\n".join(scope.scope_out)],
        ['Navigation Structure', "\n".join(scope.navigation)],
        ['Gap Analysis', "\n".join(scope.gap_analysis)],
        ['Strategic Recommendations', "\n".join(getattr(scope, 'strategic_recommendations', []))]
    ]
    df_overview = pd.DataFrame(overview_data[1:], columns=overview_data[0])
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_overview.to_excel(writer, sheet_name='Strategic Scope', index=False)
        workbook = writer.book
        sheet = workbook['Strategic Scope']
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 80
        for row in sheet.iter_rows(min_row=2, max_row=len(df_overview)+1, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrapText=True, vertical='top')
    output.seek(0)
    return output.getvalue()
